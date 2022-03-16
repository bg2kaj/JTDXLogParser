#JTDX Log Parser - BG2KAJ
#v 0.11
#2022 01 01 HNY!
#2022 01 03 v 0.11 add distance calculate summary and pack something into functions
#Feel free to modify and distribute, If you can see this line of comment.
#I create this script to dig a little bit deeper into JTDX logs so it's only for my own purpose.

import time
import json
import sys
import os
from math import sin, cos, sqrt, atan2, radians
from turtle import title
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

R = 6373.0
local_latitude=39.5
local_longitude=-115.70
dx_latitude=0
dx_longitude=0


def callsign_query(callsign,json_lib):
	callsign_length=len(callsign)
	looper=callsign_length
	query_success=0
	while(looper>0):
		callsign_tester=callsign[0:looper]
		looper-=1
		try:
			sub_data_call=json.dumps(json_lib[callsign_tester])
			sub_data_parsed_call=json.loads(sub_data_call)
		except:
			if(looper>0):
				continue
			else:
				query_success=0
				break
		else:
			query_success=1
			break

	if(query_success==1):
		return sub_data_parsed_call
	else:
		return 0
	

def freq_to_band(freq_str):
	try:
		freq_parsed=float(freq_str)
		if(freq_parsed<2):
			now_band=160
		elif(freq_parsed>3.5)and(freq_parsed<4):
			now_band=80
		elif(freq_parsed>5)and(freq_parsed<5.5):
			now_band=60
		elif(freq_parsed>7)and(freq_parsed<7.3):
			now_band=40
		elif(freq_parsed>10.1)and(freq_parsed<10.15):
			now_band=30
		elif(freq_parsed>14)and(freq_parsed<14.4):
			now_band=20
		elif(freq_parsed>18)and(freq_parsed<18.2):
			now_band=17
		elif(freq_parsed>21)and(freq_parsed<21.5):
			now_band=15
		elif(freq_parsed>24.8)and(freq_parsed<24.9):
			now_band=12
		elif(freq_parsed>28)and(freq_parsed<30):
			now_band=10
		elif(freq_parsed>50):
			now_band=6
		else:
			now_band=73

	except:
		print("Band parse failed.")
	else:
		return now_band

def calculate_distance_roughly(dx_latitude,dx_longitude,local_latitude,local_longitude):
	local_lat_ra=radians(local_latitude)
	local_long_ra=radians(local_longitude)
	dx_lat_ra=radians(dx_latitude)
	dx_long_ra=radians(dx_longitude)

	dlong=dx_long_ra-local_long_ra
	dlat=dx_lat_ra-local_lat_ra
	a = sin(dlat / 2)**2 + cos(dx_lat_ra) * cos(local_lat_ra) * sin(dlong / 2)**2
	c = 2 * atan2(sqrt(a), sqrt(1 - a))
	distance=R*c
	return distance

def draw_UI(file_size, line_average_size,line,process_date_time,process_from_to,process_result):
	print("\033[H\033[J", end="")
	print("=============================")
	print("JLP - Now processing... ")
	print("=============================")
	print("Date & Time:"+process_date_time)
	print("Info: "+process_from_to)
	print("Process result: "+process_result)
	print("=============================")
	percentage=((line*line_average_size)/(file_size))*100
	round_percentage=int(percentage//10)
	tenminusround=10-round_percentage
	percent_str="%.1f %%"%percentage
	sys.stdout.write(percent_str+" [")
	while(round_percentage>0):
		sys.stdout.write("==")
		round_percentage-=1
	while(tenminusround>0):
		sys.stdout.write("--")
		tenminusround-=1
	sys.stdout.write("]")
	print("\n=============================")
	sys.stdout.flush()


def main():
#print titles
	out=os.system("cls")
	print ("==================================================================")
	print ("JTDX Runtime Log parser By BG2KAJ v 0.11")
	print ("This is a simple script to parse JTDX 20xxxx_ALL.txt file and generate several useful log summaries.")
	print ("Only used by radio amateurs. STILL UNDER DEVELOPMENT. Use at your own risk.")
	print ("==================================================================")
	print ("Copyright informations:")
	print ("BigCty file is a work of AD1C Jim Reisert.")
	print ("datparse.py is based on work of classabbyamp, Distribute with MIT licence.")
	print ("Inspired by PyHamTools and other WSJT-X / JTDX related tools.")
#init vars
#Continent counter
	cEU_counter=0
	cAS_counter=0
	cAF_counter=0
	cNA_counter=0
	cSA_counter=0
	cOC_counter=0
	cOther_counter=0

#Band counter
	bVLF=0
	b160=0
	b80=0
	b60=0
	b40=0
	b30=0
	b20=0
	b17=0
	b15=0
	b12=0
	b10=0
	b6=0

#date counter
	day_counter=[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]
#hour counter
	hour_counter=[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]

#distance counter
	dm1k=0                #<1k km
	d1k3k=0            #1k~3k km
	d3k5k=0            #3k~5k km
	d5k7k=0            #5k~7k km
	d7k1w=0            #7k~10k km
	d1wo=0              #>10k km

	line_read_in_counter=0
	last_day=1
	last_hour=0
	now_band=40
	hour_total=0
	hour_parse_failed=0
	

	if_calculate_distance=1
	print ("==================================================================")
	print (" ")
#open config tools
	try:
		config_file=open("config.dat")
		config_line=config_file.readline()
		config_parsed=config_line.split("=")
		local_latitude=float(config_parsed[1])
		config_line=config_file.readline()
		config_parsed=config_line.split("=")
		local_longitude=float(config_parsed[1])
		print ("Config file loaded: Lat: %f Lon: %f" %(local_latitude,local_longitude))
		config_file.close()
	except:
		print("Config loading failed, Use Beijing As default QTH.")
		local_latitude=39.5
		local_longitude=-115.70


#open log file.
	try:
		input_log_file = open(str(sys.argv[1]))
	except:
		print("Input file name invalid or corrupted.")
		print("Usage: JLP.py <your log file name.txt>")
		sys.exit()

	try:
		input_if_calculate=str(sys.argv[2])
		if(input_if_calculate=="-nd"):
			if_calculate_distance=0
			print ("Command line config: No calculate of distance.")
		else:
			if_calculate_distance=1
	except:
		if_calculate_distance=1

#open result file
	log_file_size=os.path.getsize(str(sys.argv[1]))
	output_file_name="summary_of_"+input_log_file.name
	output_result_file=open(output_file_name,'w+')
	output_result_file.write("JTDX Log Parser Summary File\n")
	output_result_file.write("Result for file:"+input_log_file.name+"\n")
	process_time_start=time.time()
	localtime=time.asctime(time.localtime(time.time()))
	output_result_file.write("Task starts at :"+localtime+".\n")
	print ("Log file : "+input_log_file.name+" open successfully.")
	print ("Task starts at : "+localtime)
	output_result_file.write("\n")
	time.sleep(5)
#Create XLSX result file
	wb=Workbook()
	#open default sheet
	workbooksheet=wb.active
	#print instructions
	workbooksheet.merge_cells('A1:M1')
	nowWorkingCell=workbooksheet.cell(row=1,column=1)
	nowWorkingCell.value="JLP Analyse Report - Instructions"
	nowWorkingCell.alignment = Alignment(horizontal='center', vertical='center')
	workbooksheet.merge_cells('A2:M2')
	nowWorkingCell=workbooksheet.cell(row=2,column=1)
	nowWorkingCell.value="In this report file, every single day result is in a seperate sheet."
	nowWorkingCell.alignment = Alignment(horizontal='center', vertical='center')
	workbooksheet.merge_cells('A3:M3')
	nowWorkingCell=workbooksheet.cell(row=3,column=1)
	nowWorkingCell.value="Everyday report contains continental report, distance report and band report."
	nowWorkingCell.alignment = Alignment(horizontal='center', vertical='center')
	workbooksheet.merge_cells('A4:M4')
	nowWorkingCell=workbooksheet.cell(row=4,column=1)
	nowWorkingCell.value="This report was generated by JTDXLogParser original by BG2KAJ."
	nowWorkingCell.alignment = Alignment(horizontal='center', vertical='center')
  
  
#open database file created by datparse.py
	try:
		Lib_file=open("cty.json",'r')
		for line in Lib_file.readlines():
			json_data=json.loads(line)
	except:
		print("cty.json missing or corrupted. Run datparse.py to generate a new one from cty.dat.")
		sys.exit()

#read in a line and start parse
	line_present=input_log_file.readline()

	while line_present:
	#	print (line_present)
		line_read_in_counter=line_read_in_counter+1
		line_parsed=line_present.split(" ")

#check log entry time 
		try:
			date_time=line_parsed[0]
			date_time_parsed=date_time.split("_")
			year=int((date_time_parsed[0])[0:4])
			month=int((date_time_parsed[0])[4:6])
			day=int((date_time_parsed[0])[6:8])

	#get log entry hour
			hour=int((date_time_parsed[1])[0:2])
	#Daily report summary file write in
			if(last_day!=day):
				output_result_file.write("Day %d-%d-%d Daily detailed report:\n" %(year,month,last_day))
				output_result_file.write(str(hour_counter))
				output_result_file.write("\n")
				hour_counter=[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]
				last_day=day
				output_result_file.write("\n")
	#Create XLSX new table
				workbooksheet=wb.create_sheet("%d-%d-%d" %(year,month,day))
				#Draw full title
				workbooksheet.merge_cells('A1:H1')
				nowWorkingCell=workbooksheet.cell(row=1,column=1)
				nowWorkingCell.value=("Analyse report of %d-%d-%d " %(year,month,day))
				nowWorkingCell.alignment = Alignment(horizontal='center', vertical='center')
				#Draw continental title
				workbooksheet.merge_cells('A2:H2')
				nowWorkingCell=workbooksheet.cell(row=2,column=1)
				nowWorkingCell.value="Sort by continent"
				nowWorkingCell.alignment = Alignment(horizontal='center', vertical='center')
				#Draw subtitle for continental
				nowWorkingCell=workbooksheet.cell(row=3,column=1)
				nowWorkingCell.value="TIME/UTC"
				nowWorkingCell.alignment = Alignment(horizontal='center', vertical='center')
				nowWorkingCell=workbooksheet.cell(row=3,column=2)
				nowWorkingCell.value="AS"
				nowWorkingCell.alignment = Alignment(horizontal='center', vertical='center')
				nowWorkingCell=workbooksheet.cell(row=3,column=3)
				nowWorkingCell.value="AF"
				nowWorkingCell.alignment = Alignment(horizontal='center', vertical='center')
				nowWorkingCell=workbooksheet.cell(row=3,column=4)
				nowWorkingCell.value="EU"
				nowWorkingCell.alignment = Alignment(horizontal='center', vertical='center')
				nowWorkingCell=workbooksheet.cell(row=3,column=5)
				nowWorkingCell.value="NA"
				nowWorkingCell.alignment = Alignment(horizontal='center', vertical='center')
				nowWorkingCell=workbooksheet.cell(row=3,column=6)
				nowWorkingCell.value="OC"
				nowWorkingCell.alignment = Alignment(horizontal='center', vertical='center')
				nowWorkingCell=workbooksheet.cell(row=3,column=7)
				nowWorkingCell.value="SA"
				nowWorkingCell.alignment = Alignment(horizontal='center', vertical='center')
				nowWorkingCell=workbooksheet.cell(row=3,column=8)
				nowWorkingCell.value="OTHER"
				nowWorkingCell.alignment = Alignment(horizontal='center', vertical='center')
				#Draw row titles
				for x in range(4,28):
					nowWorkingCell=workbooksheet.cell(row=x,column=1)
					nowWorkingCell.value=x-4
					nowWorkingCell.alignment = Alignment(horizontal='center', vertical='center')
				#draw sum
				nowWorkingCell=workbooksheet.cell(row=28,column=1)
				nowWorkingCell.value="SUM"
				nowWorkingCell.alignment = Alignment(horizontal='center', vertical='center')	

				#Draw distance title
				workbooksheet.merge_cells('A29:H29')
				nowWorkingCell=workbooksheet.cell(row=29,column=1)
				nowWorkingCell.value="Sort by distance"
				nowWorkingCell.alignment = Alignment(horizontal='center', vertical='center')
				#Draw subtitle for distance
				nowWorkingCell=workbooksheet.cell(row=30,column=1)
				nowWorkingCell.value="TIME/UTC"
				nowWorkingCell.alignment = Alignment(horizontal='center', vertical='center')
				nowWorkingCell=workbooksheet.cell(row=30,column=2)
				nowWorkingCell.value="<1k"
				nowWorkingCell.alignment = Alignment(horizontal='center', vertical='center')
				nowWorkingCell=workbooksheet.cell(row=30,column=3)
				nowWorkingCell.value="1~3k"
				nowWorkingCell.alignment = Alignment(horizontal='center', vertical='center')
				nowWorkingCell=workbooksheet.cell(row=30,column=4)
				nowWorkingCell.value="3~5k"
				nowWorkingCell.alignment = Alignment(horizontal='center', vertical='center')
				nowWorkingCell=workbooksheet.cell(row=30,column=5)
				nowWorkingCell.value="5~7k"
				nowWorkingCell.alignment = Alignment(horizontal='center', vertical='center')
				nowWorkingCell=workbooksheet.cell(row=30,column=6)
				nowWorkingCell.value="7~10k"
				nowWorkingCell.alignment = Alignment(horizontal='center', vertical='center')
				nowWorkingCell=workbooksheet.cell(row=30,column=7)
				nowWorkingCell.value=">10k"
				nowWorkingCell.alignment = Alignment(horizontal='center', vertical='center')
				nowWorkingCell=workbooksheet.cell(row=30,column=8)
				nowWorkingCell.value="OTHER"
				nowWorkingCell.alignment = Alignment(horizontal='center', vertical='center')
				#Draw row titles
				for x in range(31,55):
					nowWorkingCell=workbooksheet.cell(row=x,column=1)
					nowWorkingCell.value=x-31
					nowWorkingCell.alignment = Alignment(horizontal='center', vertical='center')
				#draw sum
				nowWorkingCell=workbooksheet.cell(row=55,column=1)
				nowWorkingCell.value="SUM"
				nowWorkingCell.alignment = Alignment(horizontal='center', vertical='center')

				#Draw band title
				workbooksheet.merge_cells('A56:H56')
				nowWorkingCell=workbooksheet.cell(row=56,column=1)
				nowWorkingCell.value="Sort by band"
				nowWorkingCell.alignment = Alignment(horizontal='center', vertical='center')
				#Draw subtitle for distance
				nowWorkingCell=workbooksheet.cell(row=57,column=1)
				nowWorkingCell.value="TIME/UTC"
				nowWorkingCell.alignment = Alignment(horizontal='center', vertical='center')
				nowWorkingCell=workbooksheet.cell(row=57,column=2)
				nowWorkingCell.value="160m"
				nowWorkingCell.alignment = Alignment(horizontal='center', vertical='center')
				nowWorkingCell=workbooksheet.cell(row=57,column=3)
				nowWorkingCell.value="80m"
				nowWorkingCell.alignment = Alignment(horizontal='center', vertical='center')
				nowWorkingCell=workbooksheet.cell(row=57,column=4)
				nowWorkingCell.value="60m"
				nowWorkingCell.alignment = Alignment(horizontal='center', vertical='center')
				nowWorkingCell=workbooksheet.cell(row=57,column=5)
				nowWorkingCell.value="40m"
				nowWorkingCell.alignment = Alignment(horizontal='center', vertical='center')
				nowWorkingCell=workbooksheet.cell(row=57,column=6)
				nowWorkingCell.value="30m"
				nowWorkingCell.alignment = Alignment(horizontal='center', vertical='center')
				nowWorkingCell=workbooksheet.cell(row=57,column=7)
				nowWorkingCell.value="20m"
				nowWorkingCell.alignment = Alignment(horizontal='center', vertical='center')
				nowWorkingCell=workbooksheet.cell(row=57,column=8)
				nowWorkingCell.value="17m"
				nowWorkingCell.alignment = Alignment(horizontal='center', vertical='center')
				nowWorkingCell=workbooksheet.cell(row=57,column=9)
				nowWorkingCell.value="15m"
				nowWorkingCell.alignment = Alignment(horizontal='center', vertical='center')
				nowWorkingCell=workbooksheet.cell(row=57,column=10)
				nowWorkingCell.value="12m"
				nowWorkingCell.alignment = Alignment(horizontal='center', vertical='center')
				nowWorkingCell=workbooksheet.cell(row=57,column=11)
				nowWorkingCell.value="10m"
				nowWorkingCell.alignment = Alignment(horizontal='center', vertical='center')
				nowWorkingCell=workbooksheet.cell(row=57,column=12)
				nowWorkingCell.value="6m"
				nowWorkingCell.alignment = Alignment(horizontal='center', vertical='center')
				#Draw row titles
				for x in range(58,82):
					nowWorkingCell=workbooksheet.cell(row=x,column=1)
					nowWorkingCell.value=x-58
					nowWorkingCell.alignment = Alignment(horizontal='center', vertical='center')
				#draw sum
				nowWorkingCell=workbooksheet.cell(row=82,column=1)
				nowWorkingCell.value="SUM"
				nowWorkingCell.alignment = Alignment(horizontal='center', vertical='center')
	#An hour changed summary file output write in 
			if(last_hour!=hour):
	#hour report
				output_result_file.write("Day %d-%d-%d Hour %d detailed report:\n" %(year,month,last_day,last_hour))
				output_result_file.write("%d lines input with %d lines parse failed.\n" %(hour_total,hour_parse_failed))
				hour_total=0
				hour_parse_failed=0
	#continential report
				output_result_file.write("AF: %d AS: %d EU: %d NA: %d SA: %d OC: %d \n" %(cAF_counter,cAS_counter,cEU_counter,cNA_counter,cSA_counter,cOC_counter))

	#draw xlsx table time-continental data
				nowWorkingCell=workbooksheet.cell(row=4+last_hour,column=2)	# write AS
				nowWorkingCell.value=cAS_counter
				nowWorkingCell=workbooksheet.cell(row=4+last_hour,column=3)	# write AF
				nowWorkingCell.value=cAF_counter
				nowWorkingCell=workbooksheet.cell(row=4+last_hour,column=4)	# write EU
				nowWorkingCell.value=cEU_counter
				nowWorkingCell=workbooksheet.cell(row=4+last_hour,column=5)	# write NA
				nowWorkingCell.value=cNA_counter
				nowWorkingCell=workbooksheet.cell(row=4+last_hour,column=6)	# write OC
				nowWorkingCell.value=cOC_counter
				nowWorkingCell=workbooksheet.cell(row=4+last_hour,column=7)	# write SA
				nowWorkingCell.value=cSA_counter
				nowWorkingCell=workbooksheet.cell(row=4+last_hour,column=8)	# write OTHER
				nowWorkingCell.value=cOther_counter
	#reset counters			
				cEU_counter=0
				cAS_counter=0
				cAF_counter=0
				cNA_counter=0
				cSA_counter=0
				cOC_counter=0
				cOther_counter=0
	#band report
				output_result_file.write("160m: %d 80m: %d 60m: %d 40m: %d 30m: %d 20m: %d 17m: %d 15m: %d 12m: %d 10m: %d 6m: %d \n" %(b160,b80,b60,b40,b30,b20,b17,b15,b12,b10,b6))
	#draw xlsx table time-band data
				nowWorkingCell=workbooksheet.cell(row=58+last_hour,column=2)	# write 160
				nowWorkingCell.value=b160
				nowWorkingCell=workbooksheet.cell(row=58+last_hour,column=3)	# write 80
				nowWorkingCell.value=b80
				nowWorkingCell=workbooksheet.cell(row=58+last_hour,column=4)	# write 60
				nowWorkingCell.value=b60
				nowWorkingCell=workbooksheet.cell(row=58+last_hour,column=5)	# write 40
				nowWorkingCell.value=b40
				nowWorkingCell=workbooksheet.cell(row=58+last_hour,column=6)	# write 30
				nowWorkingCell.value=b30
				nowWorkingCell=workbooksheet.cell(row=58+last_hour,column=7)	# write 20
				nowWorkingCell.value=b20
				nowWorkingCell=workbooksheet.cell(row=58+last_hour,column=8)	# write 17
				nowWorkingCell.value=b17
				nowWorkingCell=workbooksheet.cell(row=58+last_hour,column=9)	# write 15
				nowWorkingCell.value=b15
				nowWorkingCell=workbooksheet.cell(row=58+last_hour,column=10)	# write 12
				nowWorkingCell.value=b12
				nowWorkingCell=workbooksheet.cell(row=58+last_hour,column=11)	# write 10
				nowWorkingCell.value=b10
				nowWorkingCell=workbooksheet.cell(row=58+last_hour,column=12)	# write 6
				nowWorkingCell.value=b6
	#reset counters	
				bVLF=0
				b160=0
				b80=0
				b60=0
				b40=0
				b30=0
				b20=0
				b17=0
				b15=0
				b12=0
				b10=0
				b6=0
	#distance report
				if(if_calculate_distance==1):
					output_result_file.write("0~1000 KM : %d, 1000~3000 KM:%d, 3000~5000 KM:%d, 5000~7000KM:%d, 7000~10000KM:%d, 10000~KM: %d\n\n" %(dm1k,d1k3k,d3k5k,d5k7k,d7k1w,d1wo))
	#draw xlsx table time-distance data
					nowWorkingCell=workbooksheet.cell(row=31+last_hour,column=2)	# write 1k
					nowWorkingCell.value=dm1k
					nowWorkingCell=workbooksheet.cell(row=31+last_hour,column=3)	# write 1k3k
					nowWorkingCell.value=d1k3k
					nowWorkingCell=workbooksheet.cell(row=31+last_hour,column=4)	# write 3k5k
					nowWorkingCell.value=d3k5k
					nowWorkingCell=workbooksheet.cell(row=31+last_hour,column=5)	# write 5k7k
					nowWorkingCell.value=d5k7k
					nowWorkingCell=workbooksheet.cell(row=31+last_hour,column=6)	# write 7k1w
					nowWorkingCell.value=d7k1w
					nowWorkingCell=workbooksheet.cell(row=31+last_hour,column=7)	# write 1wover
					nowWorkingCell.value=d1wo
					nowWorkingCell=workbooksheet.cell(row=31+last_hour,column=8)	# write OTHER
					nowWorkingCell.value=0
	#reset counters	
					dm1k=0                #<1k km
					d1k3k=0            #1k~3k km
					d3k5k=0            #3k~5k km
					d5k7k=0            #5k~7k km
					d7k1w=0            #7k~10k km
					d1wo=0              #>10k km
				last_hour=hour

  
  #Daily XLSX report file write in

	#parse log item
			if("CQ" in line_parsed): #CQ
				if("Transmitting" in line_parsed):
					draw_UI(log_file_size, 51,line_read_in_counter,line_parsed[0],"CQ from Myself","OK ")
				else:	#other's CQ
					hour_total+=1
					if(len(line_parsed[(line_parsed.index("CQ")+1)])==2):
						callsign_out=line_parsed[(line_parsed.index("CQ")+2)]
					else:
						callsign_out=line_parsed[(line_parsed.index("CQ")+1)]
					callsign_parse_result=callsign_query(callsign_out,json_data)
					if(callsign_parse_result!=0):
	#						print ("Continent from :"+ callsign_parse_result['continent'])
						demodulate_result="PASS"
						if(callsign_parse_result['continent']=="AS"):
							cAS_counter+=1
						elif(callsign_parse_result['continent']=="EU"):
							cEU_counter+=1
						elif(callsign_parse_result['continent']=="AF"):
							cAF_counter+=1
						elif(callsign_parse_result['continent']=="NA"):
							cNA_counter+=1
						elif(callsign_parse_result['continent']=="SA"):
							cSA_counter+=1
						elif(callsign_parse_result['continent']=="OC"):
							cOC_counter+=1
						elif(callsign_parse_result['continent']=="EU"):
							cEU_counter+=1
						if(if_calculate_distance==1):
							p2p_distance=calculate_distance_roughly(float(callsign_parse_result['lat']),float(callsign_parse_result['long']),local_latitude,local_longitude)
							if(p2p_distance>0)and(p2p_distance<=1000):
								dm1k+=1
							elif(p2p_distance>1000)and(p2p_distance<=3000):
								d1k3k+=1
							elif(p2p_distance>3000)and(p2p_distance<=5000):
								d3k5k+=1
							elif(p2p_distance>5000)and(p2p_distance<=7000):
								d5k7k+=1
							elif(p2p_distance>7000)and(p2p_distance<=10000):
								d7k1w+=1
							elif(p2p_distance>10000):
								d1wo+=1
					else:
						hour_parse_failed+=1
						demodulate_result="FAILED"
					draw_UI(log_file_size, 53,line_read_in_counter,line_parsed[0],("CQ from "+callsign_out),demodulate_result)

					if(now_band==160):
						b160+=1
					elif(now_band==80):
						b80+=1
					elif(now_band==60):
						b60+=1
					elif(now_band==40):
						b40+=1
					elif(now_band==30):
						b30+=1
					elif(now_band==20):
						b20+=1
					elif(now_band==17):
						b17+=1
					elif(now_band==15):
						b15+=1
					elif(now_band==12):
						b12+=1
					elif(now_band==10):
						b10+=1
					elif(now_band==6):
						b6+=1


					day_counter[day-1]+=1
					hour_counter[hour]+=1

			else:
				if(("MHz" in line_parsed) and (("FT8\n" in line_parsed) or ("JT65\n" in line_parsed))):	
	#			if("MHz" in line_parsed):	#change band
					draw_UI(log_file_size, 53,line_read_in_counter,line_parsed[0],("Freq. changed to "+line_parsed[2]),"OK ")
					now_band=freq_to_band(line_parsed[2])
				else:
					if("QSO" in line_parsed):
						draw_UI(log_file_size, 53,line_read_in_counter,line_parsed[0],"A QSO I made was logged.","OK ")
					else:
						if("loss" in line_parsed):
							draw_UI(log_file_size, 53,line_read_in_counter,line_parsed[0],"An audio loss was detected by JTDX.","OK ")
						else:
							try:	#other's QSO
								hour_total+=1
								callsign_out=line_parsed[(line_parsed.index("~")+2)]
								callsign_parse_result=callsign_query(callsign_out,json_data)
								if(callsign_parse_result!=0):
									demodulate_result="PASS"
									if(callsign_parse_result['continent']=="AS"):
										cAS_counter+=1
									elif(callsign_parse_result['continent']=="EU"):
										cEU_counter+=1
									elif(callsign_parse_result['continent']=="AF"):
										cAF_counter+=1
									elif(callsign_parse_result['continent']=="NA"):
										cNA_counter+=1
									elif(callsign_parse_result['continent']=="SA"):
										cSA_counter+=1
									elif(callsign_parse_result['continent']=="OC"):
										cOC_counter+=1
									elif(callsign_parse_result['continent']=="EU"):
										cEU_counter+=1
									if(if_calculate_distance==1):
										p2p_distance=calculate_distance_roughly(float(callsign_parse_result['lat']),float(callsign_parse_result['long']),local_latitude,local_longitude)
										if(p2p_distance>0)and(p2p_distance<=1000):
											dm1k+=1
										elif(p2p_distance>1000)and(p2p_distance<=3000):
											d1k3k+=1
										elif(p2p_distance>3000)and(p2p_distance<=5000):
											d3k5k+=1
										elif(p2p_distance>5000)and(p2p_distance<=7000):
											d5k7k+=1
										elif(p2p_distance>7000)and(p2p_distance<=10000):
											d7k1w+=1
										elif(p2p_distance>10000):
											d1wo+=1
								else:
									hour_parse_failed+=1
									demodulate_result="FAILED"

								draw_UI(log_file_size, 53,line_read_in_counter,line_parsed[0],("QSO: "+line_parsed[(line_parsed.index("~")+1)]+" <> "+line_parsed[(line_parsed.index("~")+2)]),demodulate_result)

								day_counter[day-1]+=1
								hour_counter[hour]+=1
								if(now_band==160):
									b160+=1
								elif(now_band==80):
									b80+=1
								elif(now_band==60):
									b60+=1
								elif(now_band==40):
									b40+=1
								elif(now_band==30):
									b30+=1
								elif(now_band==20):
									b20+=1
								elif(now_band==17):
									b17+=1
								elif(now_band==15):
									b15+=1
								elif(now_band==12):
									b12+=1
								elif(now_band==10):
									b10+=1
								elif(now_band==6):
									b6+=1
							except:
								print("Line %d parsed failed." %line_read_in_counter)

			print (" ")
			line_present=input_log_file.readline()
		except:
			print("Data parse failed.")

#after last line is over, have to output the last part of summary, print the overall summary and close files.
	output_result_file.write("Day %d %d %d Hour %d detailed report:\n" %(year,month,last_day,last_hour))
	output_result_file.write("%d lines input with %d lines parse failed.\n" %(hour_total,hour_parse_failed))
	hour_total=0
	hour_parse_failed=0
	output_result_file.write("AF:%d AS:%d EU:%d NA:%d SA:%d OC:%d \n" %(cAF_counter,cAS_counter,cEU_counter,cNA_counter,cSA_counter,cOC_counter))
#draw xlsx table time-continental data
	nowWorkingCell=workbooksheet.cell(row=4+last_hour,column=2)	# write AS
	nowWorkingCell.value=cAS_counter
	nowWorkingCell=workbooksheet.cell(row=4+last_hour,column=3)	# write AF
	nowWorkingCell.value=cAF_counter
	nowWorkingCell=workbooksheet.cell(row=4+last_hour,column=4)	# write EU
	nowWorkingCell.value=cEU_counter
	nowWorkingCell=workbooksheet.cell(row=4+last_hour,column=5)	# write NA
	nowWorkingCell.value=cNA_counter
	nowWorkingCell=workbooksheet.cell(row=4+last_hour,column=6)	# write OC
	nowWorkingCell.value=cOC_counter
	nowWorkingCell=workbooksheet.cell(row=4+last_hour,column=7)	# write SA
	nowWorkingCell.value=cSA_counter
	nowWorkingCell=workbooksheet.cell(row=4+last_hour,column=8)	# write OTHER
	nowWorkingCell.value=cOther_counter
	cEU_counter=0
	cAS_counter=0
	cAF_counter=0
	cNA_counter=0
	cSA_counter=0
	cOC_counter=0
	cOther_counter=0
	output_result_file.write("160m: %d 80m: %d 60m: %d 40m: %d 30m: %d 20m: %d 17m: %d 15m: %d 12m: %d 10m: %d 6m: %d \n" %(b160,b80,b60,b40,b30,b20,b17,b15,b12,b10,b6))
#draw xlsx table time-band data
	nowWorkingCell=workbooksheet.cell(row=58+last_hour,column=2)	# write 160
	nowWorkingCell.value=b160
	nowWorkingCell=workbooksheet.cell(row=58+last_hour,column=3)	# write 80
	nowWorkingCell.value=b80
	nowWorkingCell=workbooksheet.cell(row=58+last_hour,column=4)	# write 60
	nowWorkingCell.value=b60
	nowWorkingCell=workbooksheet.cell(row=58+last_hour,column=5)	# write 40
	nowWorkingCell.value=b40
	nowWorkingCell=workbooksheet.cell(row=58+last_hour,column=6)	# write 30
	nowWorkingCell.value=b30
	nowWorkingCell=workbooksheet.cell(row=58+last_hour,column=7)	# write 20
	nowWorkingCell.value=b20
	nowWorkingCell=workbooksheet.cell(row=58+last_hour,column=8)	# write 17
	nowWorkingCell.value=b17
	nowWorkingCell=workbooksheet.cell(row=58+last_hour,column=9)	# write 15
	nowWorkingCell.value=b15
	nowWorkingCell=workbooksheet.cell(row=58+last_hour,column=10)	# write 12
	nowWorkingCell.value=b12
	nowWorkingCell=workbooksheet.cell(row=58+last_hour,column=11)	# write 10
	nowWorkingCell.value=b10
	nowWorkingCell=workbooksheet.cell(row=58+last_hour,column=12)	# write 6
	nowWorkingCell.value=b6
#reset counters	
	bVLF=0
	b160=0
	b80=0
	b60=0
	b40=0
	b30=0
	b20=0
	b17=0
	b15=0
	b12=0
	b10=0
	b6=0
	if(if_calculate_distance==1):
		output_result_file.write("0~1000 KM : %d, 1000~3000 KM:%d, 3000~5000 KM:%d, 5000~7000KM:%d, 7000~10000KM:%d, 10000~KM: %d\n" %(dm1k,d1k3k,d3k5k,d5k7k,d7k1w,d1wo))
#draw xlsx table time-distance data
		nowWorkingCell=workbooksheet.cell(row=31+last_hour,column=2)	# write 1k
		nowWorkingCell.value=dm1k
		nowWorkingCell=workbooksheet.cell(row=31+last_hour,column=3)	# write 1k3k
		nowWorkingCell.value=d1k3k
		nowWorkingCell=workbooksheet.cell(row=31+last_hour,column=4)	# write 3k5k
		nowWorkingCell.value=d3k5k
		nowWorkingCell=workbooksheet.cell(row=31+last_hour,column=5)	# write 5k7k
		nowWorkingCell.value=d5k7k
		nowWorkingCell=workbooksheet.cell(row=31+last_hour,column=6)	# write 7k1w
		nowWorkingCell.value=d7k1w
		nowWorkingCell=workbooksheet.cell(row=31+last_hour,column=7)	# write 1wover
		nowWorkingCell.value=d1wo
		nowWorkingCell=workbooksheet.cell(row=31+last_hour,column=8)	# write OTHER
		nowWorkingCell.value=0
#reset counters	
		dm1k=0                #<1k km
		d1k3k=0            #1k~3k km
		d3k5k=0            #3k~5k km
		d5k7k=0            #5k~7k km
		d7k1w=0            #7k~10k km
		d1wo=0              #>10k km
#output a hour base output
	output_result_file.write("Day %d %d %d Hour detailed report:\n" %(year,month,last_day))
	output_result_file.write(str(hour_counter))
	output_result_file.write("\n")
	hour_counter=[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]
#output month base output
	output_result_file.write("\n")
	output_result_file.write("Month %d Total report:\n" %(month))
	output_result_file.write(str(day_counter))

#Save XLSX
	xlsxfilename=("Summary_Of_%d_%d.xlsx" %(year,month))
	wb.save(xlsxfilename)
#output last lines
	process_time_end=time.time()
	out=os.system("cls")
	print("=============================")
	print("JLP - Summary generated!  ")
	print("=============================")
	print ("File process complete, in total %d records imported.\nProcess took %d seconds to complete. Average %d entry per second.\n" %(line_read_in_counter,(process_time_end-process_time_start),line_read_in_counter/(process_time_end-process_time_start)))
	output_result_file.write("\nProcess took %d seconds to complete.\n" %(process_time_end-process_time_start))
	output_result_file.write("In total %d lines of record were imported." %line_read_in_counter)
	print("Please refer to file: "+ output_result_file.name +" and "+xlsxfilename+" for full report.\n")
#close file
	input_log_file.close()
	output_result_file.close()
	Lib_file.close()
	sys.exit()


if __name__ == '__main__':
	main()